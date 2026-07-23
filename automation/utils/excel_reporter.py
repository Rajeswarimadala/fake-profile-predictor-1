import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReporter:
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Color definitions for styling
        self.header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
        self.passed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
        self.failed_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red
        self.skipped_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Yellow
        self.summary_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Light Grey
        
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.bold_font = Font(name="Calibri", size=11, bold=True)
        self.regular_font = Font(name="Calibri", size=11)
        self.title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
        
        self.thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')

    def generate_reports(self, test_results: list):
        """Generates all 4 requested Excel files."""
        passed_tests = [t for t in test_results if t["status"] == "Passed"]
        failed_tests = [t for t in test_results if t["status"] == "Failed"]
        skipped_tests = [t for t in test_results if t["status"] == "Skipped"]
        
        # 1. Generate main Automation_Test_Report.xlsx
        self.generate_main_report(test_results, passed_tests, failed_tests, skipped_tests)
        
        # 2. Generate Passed_Test_Cases.xlsx
        self.generate_simple_filtered_report(passed_tests, "Passed", "Passed_Test_Cases.xlsx")
        
        # 3. Generate Failed_Test_Cases.xlsx
        self.generate_simple_filtered_report(failed_tests, "Failed", "Failed_Test_Cases.xlsx")
        
        # 4. Generate Execution_Summary.xlsx
        self.generate_summary_report(test_results, passed_tests, failed_tests, skipped_tests)

    def generate_main_report(self, results, passed, failed, skipped):
        file_path = os.path.join(self.output_dir, "Automation_Test_Report.xlsx")
        wb = openpyxl.Workbook()
        
        # Sheet 1: Executed Test Cases
        ws1 = wb.active
        ws1.title = "Executed Test Cases"
        self._write_test_sheet(ws1, results, "All Executed Test Cases")
        
        # Sheet 2: Passed Tests
        ws2 = wb.create_sheet(title="Passed Tests")
        self._write_test_sheet(ws2, passed, "Passed Test Cases")
        
        # Sheet 3: Failed Tests
        ws3 = wb.create_sheet(title="Failed Tests")
        self._write_test_sheet(ws3, failed, "Failed Test Cases", include_failure=True)
        
        # Sheet 4: Skipped Tests
        ws4 = wb.create_sheet(title="Skipped Tests")
        self._write_test_sheet(ws4, skipped, "Skipped Test Cases")
        
        # Sheet 5: Execution Metrics
        ws5 = wb.create_sheet(title="Execution Metrics")
        self._write_metrics_sheet(ws5, len(results), len(passed), len(failed), len(skipped))
        
        # Sheet 6: Defect Summary
        ws6 = wb.create_sheet(title="Defect Summary")
        self._write_defect_sheet(ws6, failed)
        
        # Sheet 7: Pass Rate Summary
        ws7 = wb.create_sheet(title="Pass Rate Summary")
        self._write_pass_rate_sheet(ws7, len(results), len(passed))
        
        wb.save(file_path)

    def generate_simple_filtered_report(self, tests, status_name, filename):
        file_path = os.path.join(self.output_dir, filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{status_name} Tests"
        self._write_test_sheet(ws, tests, f"{status_name} Test Cases Listing", include_failure=(status_name == "Failed"))
        wb.save(file_path)

    def generate_summary_report(self, results, passed, failed, skipped):
        file_path = os.path.join(self.output_dir, "Execution_Summary.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        self._write_metrics_sheet(ws, len(results), len(passed), len(failed), len(skipped))
        wb.save(file_path)

    def _write_test_sheet(self, ws, tests, title, include_failure=False):
        # Title Block
        ws.append([])
        ws.cell(row=2, column=2, value=title).font = self.title_font
        ws.row_dimensions[2].height = 25
        
        headers = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)"]
        if include_failure:
            headers.append("Failure Reason")
            
        start_row = 4
        start_col = 2
        
        # Write headers
        for col_idx, h in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        ws.row_dimensions[start_row].height = 20
        
        # Write rows
        current_row = start_row + 1
        for tc in tests:
            row_data = [
                tc["id"],
                tc["module"],
                tc["name"],
                tc["priority"],
                tc["status"],
                tc.get("execution_time", 0.05)
            ]
            if include_failure:
                row_data.append(tc.get("failure_reason", "Assertion Failed"))
                
            for col_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = self.regular_font
                cell.border = self.thin_border
                
                # Alignments
                if col_idx in [2, 5, 6, 7]:  # TestID, Priority, Status, ExecTime
                    cell.alignment = self.center_align
                else:
                    cell.alignment = self.left_align
                    
                # Highlight status column
                if col_idx == 6:  # Status column (TestID is 2, Module 3, Name 4, Priority 5, Status 6)
                    if val == "Passed":
                        cell.fill = self.passed_fill
                    elif val == "Failed":
                        cell.fill = self.failed_fill
                    else:
                        cell.fill = self.skipped_fill
                        
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
        # Autofit column widths
        for col in range(start_col, start_col + len(headers)):
            col_letter = get_column_letter(col)
            max_len = 0
            for r in range(start_row, current_row):
                val = ws.cell(row=r, column=col).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    def _write_metrics_sheet(self, ws, total, passed, failed, skipped):
        ws.append([])
        ws.cell(row=2, column=2, value="Execution Performance Metrics").font = self.title_font
        ws.row_dimensions[2].height = 25
        
        headers = ["Metric Category", "Count", "Percentage"]
        start_row = 4
        start_col = 2
        
        # Write headers
        for col_idx, h in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            
        pass_rate = (passed / total * 100) if total > 0 else 0
        fail_rate = (failed / total * 100) if total > 0 else 0
        skip_rate = (skipped / total * 100) if total > 0 else 0
        
        metrics = [
            ("Total Test Cases", total, "100.0%"),
            ("Passed Test Cases", passed, f"{pass_rate:.1f}%"),
            ("Failed Test Cases", failed, f"{fail_rate:.1f}%"),
            ("Skipped Test Cases", skipped, f"{skip_rate:.1f}%")
        ]
        
        current_row = start_row + 1
        for label, count, percent in metrics:
            row_data = [label, count, percent]
            for col_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = self.bold_font if col_idx == 2 else self.regular_font
                cell.border = self.thin_border
                
                if col_idx == 2:
                    cell.alignment = self.left_align
                else:
                    cell.alignment = self.center_align
                    
                # Apply color highlighting
                if label.startswith("Passed") and col_idx == 3:
                    cell.fill = self.passed_fill
                elif label.startswith("Failed") and col_idx == 3:
                    cell.fill = self.failed_fill
                elif label.startswith("Skipped") and col_idx == 3:
                    cell.fill = self.skipped_fill
                    
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
        for col in range(start_col, start_col + len(headers)):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 25

    def _write_defect_sheet(self, ws, failed):
        ws.append([])
        ws.cell(row=2, column=2, value="Defect & failure Summary").font = self.title_font
        ws.row_dimensions[2].height = 25
        
        headers = ["Defect ID", "Test Case ID", "Module", "Failure Reason", "Severity/Priority"]
        start_row = 4
        start_col = 2
        
        for col_idx, h in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            
        current_row = start_row + 1
        for idx, tc in enumerate(failed, start=1):
            row_data = [
                f"DEFECT_{idx:03d}",
                tc["id"],
                tc["module"],
                tc.get("failure_reason", "Assertion Failed"),
                tc["priority"]
            ]
            for col_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = self.regular_font
                cell.border = self.thin_border
                
                if col_idx in [2, 3, 6]:
                    cell.alignment = self.center_align
                else:
                    cell.alignment = self.left_align
                    
                if col_idx == 2:  # Defect ID is critical, color red-soft
                    cell.fill = self.failed_fill
                    cell.font = self.bold_font
                    
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
        for col in range(start_col, start_col + len(headers)):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 25

    def _write_pass_rate_sheet(self, ws, total, passed):
        ws.append([])
        ws.cell(row=2, column=2, value="Pass Rate Summary").font = self.title_font
        ws.row_dimensions[2].height = 25
        
        start_row = 4
        start_col = 2
        
        pass_rate = (passed / total * 100) if total > 0 else 0
        
        ws.cell(row=start_row, column=start_col, value="Overall Pass Percentage").font = self.header_font
        ws.cell(row=start_row, column=start_col).fill = self.header_fill
        ws.cell(row=start_row, column=start_col).alignment = self.center_align
        ws.cell(row=start_row, column=start_col).border = self.thin_border
        
        cell_rate = ws.cell(row=start_row, column=start_col + 1, value=f"{pass_rate:.2f}%")
        cell_rate.font = Font(name="Calibri", size=14, bold=True, color="006100" if pass_rate >= 95 else "9C0006")
        cell_rate.fill = self.passed_fill if pass_rate >= 95 else self.failed_fill
        cell_rate.alignment = self.center_align
        cell_rate.border = self.thin_border
        
        ws.column_dimensions[get_column_letter(start_col)].width = 28
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 20
        ws.row_dimensions[start_row].height = 30
