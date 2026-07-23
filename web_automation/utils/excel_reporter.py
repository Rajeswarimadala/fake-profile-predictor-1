import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class WebExcelReporter:
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        
        self.header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        self.passed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        self.failed_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        self.skipped_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.bold_font = Font(name="Calibri", size=11, bold=True)
        self.regular_font = Font(name="Calibri", size=11)
        self.title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
        
        self.thin_border = Border(
            left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')

    def generate_reports(self, test_results: list):
        passed = [t for t in test_results if t["status"] == "Passed"]
        failed = [t for t in test_results if t["status"] == "Failed"]
        skipped = [t for t in test_results if t["status"] == "Skipped"]
        
        # 1. Automation_Test_Report.xlsx
        wb1 = openpyxl.Workbook()
        ws1 = wb1.active
        ws1.title = "Executed Test Cases"
        self._write_test_sheet(ws1, test_results, "Executed Test Cases")
        
        ws2 = wb1.create_sheet(title="Passed Tests")
        self._write_test_sheet(ws2, passed, "Passed Test Cases")
        
        ws3 = wb1.create_sheet(title="Failed Tests")
        self._write_test_sheet(ws3, failed, "Failed Test Cases", include_failure=True)
        
        ws4 = wb1.create_sheet(title="Skipped Tests")
        self._write_test_sheet(ws4, skipped, "Skipped Test Cases")
        
        ws5 = wb1.create_sheet(title="Execution Metrics")
        self._write_metrics_sheet(ws5, len(test_results), len(passed), len(failed), len(skipped))
        
        ws6 = wb1.create_sheet(title="Defect Summary")
        self._write_defect_sheet(ws6, failed)
        
        wb1.save(os.path.join(self.output_dir, "Automation_Test_Report.xlsx"))
        
        # 2. Failed_Test_Cases.xlsx
        wb2 = openpyxl.Workbook()
        ws_f = wb2.active
        ws_f.title = "Failed Tests"
        self._write_test_sheet(ws_f, failed, "Failed Test Cases", include_failure=True)
        wb2.save(os.path.join(self.output_dir, "Failed_Test_Cases.xlsx"))

        # 3. Passed_Test_Cases.xlsx
        wb3 = openpyxl.Workbook()
        ws_p = wb3.active
        ws_p.title = "Passed Tests"
        self._write_test_sheet(ws_p, passed, "Passed Test Cases")
        wb3.save(os.path.join(self.output_dir, "Passed_Test_Cases.xlsx"))

        # 4. Summary_Report.xlsx
        wb4 = openpyxl.Workbook()
        ws_s = wb4.active
        ws_s.title = "Summary Report"
        self._write_metrics_sheet(ws_s, len(test_results), len(passed), len(failed), len(skipped))
        wb4.save(os.path.join(self.output_dir, "Summary_Report.xlsx"))

    def _write_test_sheet(self, ws, tests, title, include_failure=False):
        ws.append([])
        ws.cell(row=2, column=2, value=title).font = self.title_font
        ws.row_dimensions[2].height = 25
        
        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"]
        if include_failure:
            headers.append("Failure Reason")
            
        start_row = 4
        start_col = 2
        
        for col_idx, h in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        ws.row_dimensions[start_row].height = 20
        
        current_row = start_row + 1
        for tc in tests:
            row_data = [
                tc["id"], tc["module"], tc["name"], tc["status"],
                tc.get("execution_time", 0.05), tc["priority"]
            ]
            if include_failure:
                row_data.append(tc.get("failure_reason", "Assertion Error"))
                
            for col_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = self.regular_font
                cell.border = self.thin_border
                cell.alignment = self.center_align if col_idx in [2, 5, 6, 7] else self.left_align
                
                if col_idx == 5: # Status column index in table (col_idx=5 is 5th header: Status)
                    if val == "Passed":
                        cell.fill = self.passed_fill
                    elif val == "Failed":
                        cell.fill = self.failed_fill
                    else:
                        cell.fill = self.skipped_fill
                        
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
        for col in range(start_col, start_col + len(headers)):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 24

    def _write_metrics_sheet(self, ws, total, passed, failed, skipped):
        ws.append([])
        ws.cell(row=2, column=2, value="Execution Summary Metrics").font = self.title_font
        ws.row_dimensions[2].height = 25
        
        headers = ["Metric Category", "Count", "Percentage"]
        start_row = 4
        start_col = 2
        
        for col_idx, h in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            
        pass_rate = (passed / total * 100) if total > 0 else 0
        fail_rate = (failed / total * 100) if total > 0 else 0
        skip_rate = (skipped / total * 100) if total > 0 else 0
        
        metrics = [
            ("Total Test Cases", total, "100.0%"),
            ("Passed Test Cases", passed, f"{pass_rate:.1f}%"),
            ("Failed Test Cases", failed, f"{fail_rate:.1f}%"),
            ("Skipped Test Cases", skipped, f"{skip_rate:.1f}%")
        ]
        
        current_row = start_row + 1
        for label, count, percent in metrics:
            row_data = [label, count, percent]
            for col_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = self.bold_font if col_idx == 2 else self.regular_font
                cell.border = self.thin_border
                cell.alignment = self.left_align if col_idx == 2 else self.center_align
            current_row += 1
            
        for col in range(start_col, start_col + len(headers)):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _write_defect_sheet(self, ws, failed):
        ws.append([])
        ws.cell(row=2, column=2, value="Defect Summary").font = self.title_font
        ws.row_dimensions[2].height = 25
        
        headers = ["Defect ID", "Test Case ID", "Module", "Failure Reason", "Priority"]
        start_row = 4
        start_col = 2
        
        for col_idx, h in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            
        current_row = start_row + 1
        for idx, tc in enumerate(failed, start=1):
            row_data = [f"DEFECT_{idx:03d}", tc["id"], tc["module"], tc.get("failure_reason", "Assertion Failed"), tc["priority"]]
            for col_idx, val in enumerate(row_data, start=start_col):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = self.regular_font
                cell.border = self.thin_border
                if col_idx == 2:
                    cell.fill = self.failed_fill
                    cell.font = self.bold_font
            current_row += 1
            
        for col in range(start_col, start_col + len(headers)):
            ws.column_dimensions[get_column_letter(col)].width = 25
